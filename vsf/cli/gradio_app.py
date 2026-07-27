# Copyright (c) Opendatalab. All rights reserved.

import asyncio
import html as html_lib
import httpx
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import uuid
import zipfile
from contextlib import asynccontextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable
from urllib.parse import quote

import click
import gradio as gr
from gradio_pdf import PDF
from loguru import logger

os.environ["TORCH_CUDNN_V8_API_DISABLED"] = "1"
# Implementation detail.
_gradio_major_version = int(gr.__version__.split('.')[0])
IS_GRADIO_6 = _gradio_major_version >= 6

log_level = os.getenv("VSF_LOG_LEVEL", "INFO").upper()
logger.remove()  # Remove invalid or unnecessary data.
logger.add(sys.stderr, level=log_level)  # Add the value to the result.

from vsf.cli.common import (
    image_suffixes,
    normalize_task_stem,
    office_suffixes,
    pdf_suffixes,
    read_fn,
)
from vsf.cli import api_client as _api_client
from vsf.cli.backend_options import (
    DEFAULT_BACKEND,
    DEFAULT_HYBRID_EFFORT,
    HTTP_CLIENT_BACKEND_CHOICES,
    LOCAL_BACKEND_CHOICES,
)
from vsf.cli.client_side_output import regenerate_client_side_outputs
from vsf.cli.output_paths import resolve_parse_dir
from vsf.cli.vlm_preload import resolve_gradio_local_api_cli_args
from vsf.cli.visualization import VisualizationJob, run_visualization_job
from vsf.idp.schemas import DOCUMENT_SCHEMAS

_gradio_local_api_server = _api_client.ReusableLocalAPIServer()


@dataclass(frozen=True)
class GradioConcurrencyWaitSnapshot:
    limit: int
    active: int
    waiting: int
    ahead: int


@dataclass
class _LimiterState:
    semaphore: asyncio.Semaphore
    active: int = 0
    waiters: list[object] = field(default_factory=list)


class GradioRequestConcurrencyLimiter:
    def __init__(self):
        self._lock = threading.Lock()
        self._states: dict[int, _LimiterState] = {}

    def _get_state(self, limit: int):
        if limit <= 0:
            return None
        with self._lock:
            state = self._states.get(limit)
            if state is None:
                state = _LimiterState(semaphore=asyncio.Semaphore(limit))
                self._states[limit] = state
            return state

    def _build_wait_snapshot(
        self,
        state: _LimiterState,
        limit: int,
        wait_token: object,
    ) -> GradioConcurrencyWaitSnapshot | None:
        if wait_token not in state.waiters:
            return None

        return GradioConcurrencyWaitSnapshot(
            limit=limit,
            active=state.active,
            waiting=len(state.waiters),
            ahead=state.waiters.index(wait_token),
        )

    def _remove_waiter(self, state: _LimiterState, wait_token: object) -> None:
        if wait_token in state.waiters:
            state.waiters.remove(wait_token)

    async def _cleanup_acquire_interruption(
        self,
        state: _LimiterState,
        acquire_task: asyncio.Task[bool],
        wait_token: object,
        should_wait: bool,
    ) -> None:
        if not acquire_task.done():
            acquire_task.cancel()
            await asyncio.gather(acquire_task, return_exceptions=True)
        elif not acquire_task.cancelled():
            try:
                acquired = acquire_task.result()
            except Exception:
                acquired = False
            if acquired:
                state.semaphore.release()

        if should_wait:
            with self._lock:
                self._remove_waiter(state, wait_token)

    @asynccontextmanager
    async def acquire(
        self,
        limit: int,
        on_wait: Callable[[GradioConcurrencyWaitSnapshot], None] | None = None,
    ):
        state = self._get_state(limit)
        if state is None:
            yield
            return

        wait_token = object()
        should_wait = False
        snapshot = None
        with self._lock:
            if state.active >= limit or state.waiters:
                state.waiters.append(wait_token)
                should_wait = True
                snapshot = self._build_wait_snapshot(state, limit, wait_token)

        acquire_task: asyncio.Task[bool] = asyncio.create_task(state.semaphore.acquire())
        last_wait_ahead = None
        if should_wait and on_wait is not None and snapshot is not None:
            on_wait(snapshot)
            last_wait_ahead = snapshot.ahead

        try:
            if should_wait:
                while True:
                    done, _ = await asyncio.wait(
                        {acquire_task},
                        timeout=STATUS_TIMER_INTERVAL_SECONDS,
                    )
                    if acquire_task in done:
                        acquire_task.result()
                        break

                    if on_wait is None:
                        continue

                    with self._lock:
                        snapshot = self._build_wait_snapshot(state, limit, wait_token)

                    if snapshot is None or snapshot.ahead == last_wait_ahead:
                        continue

                    on_wait(snapshot)
                    last_wait_ahead = snapshot.ahead
            else:
                await acquire_task
        except BaseException:
            await self._cleanup_acquire_interruption(
                state=state,
                acquire_task=acquire_task,
                wait_token=wait_token,
                should_wait=should_wait,
            )
            raise

        with self._lock:
            if should_wait:
                self._remove_waiter(state, wait_token)
            state.active += 1
        try:
            yield
        finally:
            with self._lock:
                state.active = max(0, state.active - 1)
            state.semaphore.release()


_gradio_request_concurrency_limiter = GradioRequestConcurrencyLimiter()

STATUS_BOX_AUTOSCROLL_JS = """
(value) => {
    const scrollToBottom = () => {
        const textarea = document.querySelector(".convert-status-box textarea");
        if (!textarea) {
            return;
        }
        textarea.scrollTop = textarea.scrollHeight;
    };

    requestAnimationFrame(() => {
        scrollToBottom();
        requestAnimationFrame(scrollToBottom);
    });

    return [];
}
"""

STATUS_TIMER_INTERVAL_SECONDS = 0.1
STATUS_QUEUE_ANIMATION_INTERVAL_SECONDS = 1.0
STATUS_QUEUE_ANIMATION_MAX_DOTS = 10

STATUS_PREPARING_REQUEST = "Đang chuẩn bị yêu cầu..."
STATUS_CHECKING_SERVER = "Đang kiểm tra dịch vụ..."
STATUS_SUBMITTING_TASK = "Đang gửi tác vụ..."
STATUS_DOWNLOADING_RESULT = "Đã xử lý xong, đang tải kết quả..."
STATUS_PROCESSING_OUTPUT = "Đang tạo tệp kết quả..."
STATUS_COMPLETED = "Hoàn thành"
STATUS_QUEUED_ON_SERVER = "Đang chờ trên máy chủ"
STATUS_PROCESSING_ON_SERVER = "Máy chủ đang xử lý"
STATUS_QUEUED_LOCALLY_PREFIX = "Đang chờ cục bộ:"

BACKEND_CHOICE_DEFINITIONS = list(LOCAL_BACKEND_CHOICES)
HTTP_CLIENT_BACKEND_CHOICE_DEFINITIONS = list(HTTP_CLIENT_BACKEND_CHOICES)
STATUS_STEP_DEFINITIONS = [
    ("status_step_prepare", STATUS_PREPARING_REQUEST),
    ("status_step_check", STATUS_CHECKING_SERVER),
    ("status_step_submit", STATUS_SUBMITTING_TASK),
    ("status_step_queue", STATUS_QUEUED_ON_SERVER),
    ("status_step_process", STATUS_PROCESSING_ON_SERVER),
    ("status_step_download", STATUS_DOWNLOADING_RESULT),
    ("status_step_outputs", STATUS_PROCESSING_OUTPUT),
    ("status_step_done", STATUS_COMPLETED),
]


def normalize_vsf_locale(locale):
    """Implementation detail."""
    normalized = str(locale or "").strip().lower()
    if normalized.startswith("zh"):
        return "zh"
    return "en"


def resolve_i18n_text(i18n, key, locale=None):
    """Extract the required value."""
    if i18n is None:
        return key
    translations = getattr(i18n, "translations", None)
    if translations:
        preferred_locale = normalize_vsf_locale(
            locale or os.getenv("VSF_GRADIO_DEFAULT_LOCALE", "zh")
        )
        preferred_text = translations.get(preferred_locale, {}).get(key)
        if preferred_text is not None:
            return preferred_text
        fallback_text = translations.get("en", {}).get(key)
        if fallback_text is not None:
            return fallback_text
        return key
    return i18n(key)


def translate_ui(i18n, key, locale=None):
    """Extract the required value."""
    return resolve_i18n_text(i18n, key, locale)


def resolve_request_locale(request):
    """Process the service request."""
    headers = getattr(request, "headers", None) or {}
    if not hasattr(headers, "get"):
        return None

    accept_language = headers.get("accept-language") or headers.get("Accept-Language")
    if not accept_language:
        return None

    language_candidates = []
    for order, raw_item in enumerate(str(accept_language).split(",")):
        parts = [part.strip() for part in raw_item.split(";") if part.strip()]
        if not parts:
            continue
        quality = 1.0
        for parameter in parts[1:]:
            name, separator, value = parameter.partition("=")
            if separator and name.strip().lower() == "q":
                try:
                    quality = float(value)
                except ValueError:
                    quality = 0.0
                break
        language_candidates.append((-quality, order, parts[0]))

    if not language_candidates:
        return None
    _, _, preferred_language = min(language_candidates)
    return normalize_vsf_locale(preferred_language)


def build_client_i18n_attrs(i18n, key):
    """Prepare the output value."""
    attrs = [f'data-vsf-i18n-key="{html_lib.escape(key, quote=True)}"']
    for locale in ("en", "zh"):
        text = resolve_i18n_text(i18n, key, locale)
        attrs.append(f'data-vsf-i18n-{locale}="{html_lib.escape(text, quote=True)}"')
    return " ".join(attrs)


def render_client_i18n_text(i18n, key, locale=None):
    """Build the required output."""
    return (
        f"<span {build_client_i18n_attrs(i18n, key)}>"
        f"{html_lib.escape(translate_ui(i18n, key, locale))}"
        "</span>"
    )


def build_backend_choices(http_client_enable, i18n):
    """Build the required output."""
    choices = list(BACKEND_CHOICE_DEFINITIONS)
    if http_client_enable:
        choices.extend(HTTP_CLIENT_BACKEND_CHOICE_DEFINITIONS)
    return choices


def is_http_client_backend(backend_choice):
    """Validate the current value."""
    return isinstance(backend_choice, str) and backend_choice.endswith("-http-client")


def select_backend_info_key(backend_choice):
    """Parse the input data."""
    if not isinstance(backend_choice, str):
        return "backend_info_default"
    if backend_choice.startswith("vlm"):
        return "backend_info_vlm"
    if backend_choice == "pipeline":
        return "backend_info_pipeline"
    if backend_choice.startswith("hybrid"):
        return "backend_info_hybrid"
    return "backend_info_default"


def select_force_ocr_info_key(backend_choice: object) -> str:
    """Parse the input data."""
    if isinstance(backend_choice, str) and backend_choice.startswith("hybrid"):
        return "force_ocr_info_hybrid"
    return "force_ocr_info"


def is_effort_option_visible(backend_choice):
    """Validate the current value."""
    return isinstance(backend_choice, str) and backend_choice.startswith("hybrid")


def resolve_status_step_index(status_lines):
    """Implementation detail."""
    if not status_lines:
        return -1, False
    if status_lines[-1].startswith("Thất bại:"):
        return len(STATUS_STEP_DEFINITIONS) - 1, True
    if any(line.startswith(STATUS_COMPLETED) for line in status_lines):
        return len(STATUS_STEP_DEFINITIONS) - 1, False

    for index in range(len(STATUS_STEP_DEFINITIONS) - 1, -1, -1):
        _, marker = STATUS_STEP_DEFINITIONS[index]
        if marker == STATUS_QUEUED_ON_SERVER:
            if any(StatusPanelState.is_queue_message(line) for line in status_lines):
                return index, False
            continue
        if any(line.startswith(marker) for line in status_lines):
            return index, False
    return 0, False


def render_status_steps_html(status_text, i18n, locale=None):
    """Implementation detail."""
    status_lines = [line for line in str(status_text or "").splitlines() if line]
    current_index, is_failed = resolve_status_step_index(status_lines)
    latest_status = (
        status_lines[-1]
        if status_lines
        else render_client_i18n_text(i18n, "status_idle_hint", locale)
    )

    step_items = []
    for index, (label_key, _) in enumerate(STATUS_STEP_DEFINITIONS):
        classes = ["status-step"]
        if is_failed and index == current_index:
            classes.extend(["is-active", "is-error"])
            label = render_client_i18n_text(i18n, "status_step_failed", locale)
        elif index < current_index or (
            current_index == len(STATUS_STEP_DEFINITIONS) - 1 and not is_failed
        ):
            classes.append("is-done")
            label = render_client_i18n_text(i18n, label_key, locale)
        elif index == current_index:
            classes.append("is-active")
            label = render_client_i18n_text(i18n, label_key, locale)
        else:
            classes.append("is-pending")
            label = render_client_i18n_text(i18n, label_key, locale)
        step_items.append(
            f'<div class="{" ".join(classes)}">'
            f'<span class="status-dot"></span>'
            f'<span class="status-label">{label}</span>'
            "</div>"
        )

    title_key = "status_idle_title" if not status_lines else "status_latest"
    return (
        '<div class="status-steps-panel">'
        f'<div class="status-panel-title">'
        f'{render_client_i18n_text(i18n, title_key, locale)}'
        f'</div>'
        f'<div class="status-steps-list">{"".join(step_items)}</div>'
        f'<div class="status-latest">'
        f'{latest_status if not status_lines else html_lib.escape(latest_status)}'
        f'</div>'
        "</div>"
    )


RESOURCE_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'resources')


def load_resource_text(resource_name):
    """Extract the required value."""
    resource_path = os.path.join(RESOURCE_DIR, resource_name)
    with open(resource_path, mode='r', encoding='utf-8') as resource_file:
        return resource_file.read()


APP_CSS = load_resource_text('gradio_app.css')
APP_JS = load_resource_text('gradio_app.js')

# Process text content.
APP_HEAD = f"""
<script>
(() => {{
    const installVSFAdvancedPopover = {APP_JS};
    if (document.readyState === "loading") {{
        document.addEventListener("DOMContentLoaded", installVSFAdvancedPopover, {{ once: true }});
    }} else {{
        installVSFAdvancedPopover();
    }}
}})();
</script>
"""


@dataclass
class StatusPanelState:
    lines: list[str] = field(default_factory=list)
    processing_index: int | None = None
    processing_started_at: float | None = None
    last_processing_elapsed_seconds: float | None = None
    queue_index: int | None = None
    queue_started_at: float | None = None
    queue_base_message: str | None = None

    def append(self, message: str) -> bool:
        if not message:
            return False

        if self.is_queue_message(message):
            self.finalize_processing()
            return self.update_queue(message)

        if message == STATUS_PROCESSING_ON_SERVER:
            self.finalize_queue()
            return self.start_processing()

        self.finalize_processing()
        self.finalize_queue()
        if message == STATUS_COMPLETED:
            message = format_completed_status(self.last_processing_elapsed_seconds)
        if not self.lines or self.lines[-1] != message:
            self.lines.append(message)
            return True
        return False

    def start_processing(self) -> bool:
        if self.processing_started_at is not None:
            return self.tick_processing()

        self.processing_started_at = time.monotonic()
        self.last_processing_elapsed_seconds = 0.0
        self.processing_index = len(self.lines)
        self.lines.append(format_processing_status(0.0))
        return True

    def tick_processing(self) -> bool:
        if self.processing_started_at is None or self.processing_index is None:
            return False

        elapsed_seconds = max(0.0, time.monotonic() - self.processing_started_at)
        self.last_processing_elapsed_seconds = elapsed_seconds
        updated = format_processing_status(elapsed_seconds)
        if self.lines[self.processing_index] != updated:
            self.lines[self.processing_index] = updated
            return True
        return False

    def finalize_processing(self) -> bool:
        if self.processing_started_at is None or self.processing_index is None:
            return False

        self.tick_processing()
        self.processing_started_at = None
        self.processing_index = None
        return True

    def update_queue(self, message: str) -> bool:
        if (
            self.queue_index is None
            or self.queue_started_at is None
            or self.queue_base_message is None
        ):
            self.queue_started_at = time.monotonic()
            self.queue_index = len(self.lines)
            self.queue_base_message = message
            self.lines.append(format_queue_status(message, 0.0))
            return True

        self.queue_base_message = message
        updated = format_queue_status(
            message,
            max(0.0, time.monotonic() - self.queue_started_at),
        )
        if self.lines[self.queue_index] != updated:
            self.lines[self.queue_index] = updated
            return True
        return False

    def tick_queue(self) -> bool:
        if (
            self.queue_index is None
            or self.queue_started_at is None
            or self.queue_base_message is None
        ):
            return False

        updated = format_queue_status(
            self.queue_base_message,
            max(0.0, time.monotonic() - self.queue_started_at),
        )
        if self.lines[self.queue_index] != updated:
            self.lines[self.queue_index] = updated
            return True
        return False

    def finalize_queue(self) -> bool:
        if (
            self.queue_index is None
            or self.queue_started_at is None
            or self.queue_base_message is None
        ):
            return False

        self.tick_queue()
        self.queue_index = None
        self.queue_started_at = None
        self.queue_base_message = None
        return True

    def tick(self) -> bool:
        if self.is_processing:
            return self.tick_processing()
        if self.is_queueing:
            return self.tick_queue()
        return False

    @property
    def is_processing(self) -> bool:
        return self.processing_started_at is not None

    @property
    def is_queueing(self) -> bool:
        return self.queue_started_at is not None

    @property
    def animation_interval_seconds(self) -> float | None:
        if self.is_processing:
            return STATUS_TIMER_INTERVAL_SECONDS
        if self.is_queueing:
            return STATUS_QUEUE_ANIMATION_INTERVAL_SECONDS
        return None

    @staticmethod
    def is_queue_message(message: str) -> bool:
        return (
            message.startswith(STATUS_QUEUED_LOCALLY_PREFIX)
            or message.startswith(STATUS_QUEUED_ON_SERVER)
        )

    def render(self) -> str:
        return "\n".join(self.lines)


def format_failed_status(error: Exception | str) -> str:
    return f"Thất bại: {error}"


def format_processing_status(elapsed_seconds: float) -> str:
    return f"{STATUS_PROCESSING_ON_SERVER} ({elapsed_seconds:.1f}s)"


def format_completed_status(elapsed_seconds: float | None) -> str:
    """Parse the input data."""
    if elapsed_seconds is None:
        return STATUS_COMPLETED
    return f"{STATUS_COMPLETED} ({elapsed_seconds:.1f}s)"


def format_queue_status(base_message: str, elapsed_seconds: float) -> str:
    dots = "." * (
        (int(max(0.0, elapsed_seconds)) % STATUS_QUEUE_ANIMATION_MAX_DOTS) + 1
    )
    return f"{base_message}{dots}"


def format_concurrency_wait_message(snapshot: GradioConcurrencyWaitSnapshot) -> str:
    return f"{STATUS_QUEUED_LOCALLY_PREFIX} còn {snapshot.ahead} tác vụ phía trước"


def format_remote_status_message(
    status_snapshot: _api_client.TaskStatusSnapshot | str,
) -> str:
    if isinstance(status_snapshot, _api_client.TaskStatusSnapshot):
        status = status_snapshot.status
        queued_ahead = status_snapshot.queued_ahead
    else:
        status = status_snapshot
        queued_ahead = None

    if status == "pending":
        if queued_ahead is not None:
            return f"{STATUS_QUEUED_ON_SERVER}: còn {queued_ahead} tác vụ phía trước"
        return STATUS_QUEUED_ON_SERVER
    if status == "processing":
        return STATUS_PROCESSING_ON_SERVER
    if status == "completed":
        return STATUS_COMPLETED
    if status == "failed":
        return format_failed_status("tác vụ trên máy chủ gặp lỗi")
    return f"Trạng thái tác vụ: {status}"


def compress_directory_to_zip(directory_path, output_zip_path):
    """Process the file path.

    Process the file path.
    Process the file path.
    """
    try:
        with zipfile.ZipFile(output_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:

            # Iterate over the available items.
            for root, dirs, files in os.walk(directory_path):
                for file in files:
                    # Build the required output.
                    file_path = os.path.join(root, file)
                    # Calculate the result.
                    arcname = os.path.relpath(file_path, directory_path)
                    # Add the value to the result.
                    zipf.write(file_path, arcname)
        return 0
    except Exception as e:
        logger.exception(e)
        return -1


GRADIO_PREVIEW_IMAGE_SUFFIXES = {
    ".jpg",
    ".jpeg",
    ".png",
    ".gif",
    ".webp",
    ".svg",
}
GRADIO_PREVIEW_EXTERNAL_SRC_RE = re.compile(r"^[a-zA-Z][a-zA-Z0-9+.-]*:")


def _resolve_gradio_preview_image_path(src, image_dir_path):
    """Parse the input data."""
    image_src = str(src).strip()
    if not image_src:
        return None
    if Path(image_src).suffix.lower() not in GRADIO_PREVIEW_IMAGE_SUFFIXES:
        return None

    resolved_path = (Path(image_dir_path) / image_src).resolve(strict=False)
    if not resolved_path.is_file():
        logger.warning(f"Skip missing Gradio preview image: {resolved_path}")
        return None
    return str(resolved_path)


def replace_image_with_gradio_file_urls(markdown_text, image_dir_path):
    """Process image content."""
    if not isinstance(markdown_text, str) or not image_dir_path:
        return markdown_text

    def _path_to_public_url(image_src):
        image_path = _resolve_gradio_preview_image_path(image_src, image_dir_path)
        if not image_path:
            return None
        return f"/gradio_api/file={quote(image_path, safe='/:')}"

    # Match the expected pattern.
    def replace_md(match):
        alt_text = match.group("alt")
        image_src = match.group("src")
        public_url = _path_to_public_url(image_src)
        if public_url:
            return f"![{alt_text}]({public_url})"
        return match.group(0)

    result = re.sub(
        r"!\[(?P<alt>[^\]]*)\]\((?P<src>[^)]+)\)",
        replace_md,
        markdown_text,
    )

    # Remove invalid or unnecessary data.
    def replace_html_src(match):
        prefix = match.group("prefix")
        quote_char = match.group("quote")
        image_src = match.group("src")
        public_url = _path_to_public_url(image_src)
        if public_url:
            return f"{prefix}{quote_char}{public_url}{quote_char}"
        return match.group(0)

    result = re.sub(
        r"(?P<prefix><img\b[^>]*?\bsrc\s*=\s*)(?P<quote>[\"'])(?P<src>[^\"']+)(?P=quote)",
        replace_html_src,
        result,
        flags=re.IGNORECASE,
    )

    return result


def read_gradio_content_list_json(local_md_dir, file_name):
    """Parse the input data."""
    content_list_path = Path(local_md_dir) / f"{file_name}_content_list.json"
    if not content_list_path.is_file():
        logger.warning(
            f"Content list JSON not found for Gradio preview: {content_list_path}"
        )
        return ""
    try:
        return content_list_path.read_text(encoding="utf-8")
    except Exception as exc:
        logger.warning(
            f"Failed to read Gradio content list JSON: {content_list_path}, error={exc}"
        )
        return ""


def read_gradio_idp_json(local_md_dir, file_name):
    """Read the optional HR IDP result for preview."""

    idp_path = Path(local_md_dir) / f"{file_name}_idp.json"
    if not idp_path.is_file():
        return ""
    try:
        return idp_path.read_text(encoding="utf-8")
    except Exception as exc:
        logger.warning(f"Failed to read Gradio IDP JSON: {idp_path}, error={exc}")
        return ""


def _escape_latex_html_chars_for_gradio(content):
    """Parse the input data."""
    return content.replace("<", "&lt;").replace(">", "&gt;")


def escape_latex_blocks_for_gradio_preview(markdown_text, latex_delimiters):
    """Process formula content."""
    if not markdown_text or not latex_delimiters:
        return markdown_text

    delimiter_pairs = []
    for delimiter in latex_delimiters:
        left = delimiter.get("left")
        right = delimiter.get("right")
        if left and right:
            delimiter_pairs.append((left, right))
    delimiter_pairs.sort(key=lambda pair: len(pair[0]), reverse=True)
    if not delimiter_pairs:
        return markdown_text

    result = []
    position = 0
    text_length = len(markdown_text)
    while position < text_length:
        matched_pair = None
        for left, right in delimiter_pairs:
            if markdown_text.startswith(left, position):
                matched_pair = (left, right)
                break

        if matched_pair is None:
            result.append(markdown_text[position])
            position += 1
            continue

        left, right = matched_pair
        content_start = position + len(left)
        content_end = markdown_text.find(right, content_start)
        if content_end == -1:
            # Process formula content.
            result.append(markdown_text[position])
            position += 1
            continue

        result.append(left)
        result.append(
            _escape_latex_html_chars_for_gradio(markdown_text[content_start:content_end])
        )
        result.append(right)
        position = content_end + len(right)

    return "".join(result)


def prepare_markdown_for_gradio_preview(markdown_text, latex_delimiters):
    """Process text content."""
    if not isinstance(markdown_text, str):
        return markdown_text
    return escape_latex_blocks_for_gradio_preview(markdown_text, latex_delimiters)


def normalize_language(language):
    if '(' in language and ')' in language:
        return language.split('(')[0].strip()
    return language


def resolve_parse_method(file_path, is_ocr, backend):
    file_suffix = Path(file_path).suffix.lower().lstrip('.')
    if file_suffix in office_suffixes:
        return "auto"
    if backend.startswith("vlm"):
        return "auto"
    return "ocr" if is_ocr else "auto"


def is_image_analysis_option_visible(backend, effort=DEFAULT_HYBRID_EFFORT):
    """Validate the current value."""
    if not isinstance(backend, str):
        return False
    if backend.startswith("vlm"):
        return True
    if backend.startswith("hybrid"):
        return effort == "high"
    return False


def is_ocr_language_option_visible(backend: object) -> bool:
    """Validate the current value."""
    return backend == "pipeline"


def is_force_ocr_option_visible(backend: object) -> bool:
    """Validate the current value."""
    if not isinstance(backend, str):
        return False
    return backend == "pipeline" or backend.startswith("hybrid")


def frontend_managed_initial_visibility(is_visible: bool):
    """Convert the value to the required format."""
    return True if is_visible else "hidden"


def should_use_client_side_output_generation(client_side_output_generation):
    """Validate the current value."""
    return client_side_output_generation


def create_gradio_run_paths(file_path, output_root="./output"):
    run_id = f"{time.strftime('%y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}_{safe_stem(Path(file_path).stem)}"
    run_root = Path(output_root) / "gradio" / run_id
    extract_root = run_root / "result"
    archive_zip_path = run_root / f"{safe_stem(Path(file_path).stem)}.zip"
    return run_root, extract_root, archive_zip_path


def build_gradio_allowed_paths(output_root="./output"):
    """Build the required output."""
    allowed_paths = []
    for item in os.environ.get("GRADIO_ALLOWED_PATHS", "").split(","):
        item = item.strip()
        if item:
            allowed_paths.append(item)

    output_path = str(Path(output_root).resolve())
    if output_path not in allowed_paths:
        allowed_paths.append(output_path)
    return allowed_paths


def build_gradio_upload_name(file_path):
    path = Path(file_path)
    return f"{normalize_task_stem(path.stem)}{path.suffix}"


def resolve_result_file_name(submit_response, extract_root, file_path):
    if submit_response.file_names:
        return submit_response.file_names[0]

    candidate_dirs = sorted(path.name for path in Path(extract_root).iterdir() if path.is_dir())
    if len(candidate_dirs) == 1:
        return candidate_dirs[0]
    return normalize_task_stem(Path(file_path).stem)


async def resolve_server_health(http_client, api_url):
    if api_url:
        return await _api_client.fetch_server_health(
            http_client,
            _api_client.normalize_base_url(api_url),
        )

    local_server, started_now = _gradio_local_api_server.ensure_started()
    if started_now:
        logger.info(f"Started local vsf-api at {local_server.base_url}")
    return await _api_client.wait_for_local_api_ready(http_client, local_server)


async def ensure_local_api_ready_for_gradio_startup(
    timeout_seconds: float = _api_client.LOCAL_API_STARTUP_TIMEOUT_SECONDS,
):
    local_server, started_now = _gradio_local_api_server.ensure_started()
    if started_now:
        logger.info(f"Started local vsf-api at {local_server.base_url}")

    async with httpx.AsyncClient(
        timeout=_api_client.build_http_timeout(),
        follow_redirects=True,
    ) as http_client:
        return await _api_client.wait_for_local_api_ready(
            http_client,
            local_server,
            timeout_seconds=timeout_seconds,
        )


def maybe_prepare_local_api_for_gradio_startup(
    *,
    api_url: str | None,
    enable_vlm_preload: bool,
):
    if api_url is not None or not enable_vlm_preload:
        return None

    try:
        return asyncio.run(ensure_local_api_ready_for_gradio_startup())
    except Exception:
        _gradio_local_api_server.stop()
        raise


def resolve_gradio_max_concurrent_requests(api_url, server_health):
    if api_url is None:
        return server_health.max_concurrent_requests

    return _api_client.resolve_effective_max_concurrent_requests(
        local_max=_api_client.read_max_concurrent_requests(
            default=_api_client.DEFAULT_MAX_CONCURRENT_REQUESTS
        ),
        server_max=server_health.max_concurrent_requests,
    )


def maybe_generate_local_preview(extract_root, file_name, file_suffix, backend, parse_method):
    if file_suffix in office_suffixes:
        return None

    parse_dir = resolve_parse_dir(
        extract_root,
        file_name,
        backend,
        parse_method,
        allow_office_fallback=True,
    )
    visualization_job = VisualizationJob(
        document_stem=file_name,
        backend=backend,
        parse_method=parse_method,
        parse_dir=parse_dir,
        draw_span=backend.startswith("pipeline"),
    )
    result = run_visualization_job(visualization_job)
    if result.status != "finished":
        logger.warning(
            f"Skipping visualization for {visualization_job.document_stem}: {result.message}"
        )
    return resolve_preview_pdf_path(parse_dir, file_name)


async def _run_to_markdown_job(
    file_path,
    end_pages=10,
    is_ocr=False,
    formula_enable=True,
    table_enable=True,
    image_analysis=True,
    effort=DEFAULT_HYBRID_EFFORT,
    language="ch",
    backend="pipeline",
    url=None,
    api_url=None,
    client_side_output_generation=False,
    enable_idp=False,
    idp_document_type="auto",
    status_callback: Callable[[str], None] | None = None,
):
    if file_path is None:
        return "", "", "", None, None

    def emit_status(message: str) -> None:
        if status_callback is not None:
            status_callback(message)

    normalized_language = normalize_language(language)
    file_path = str(file_path)
    file_suffix = Path(file_path).suffix.lower().lstrip('.')
    use_client_side_output_generation = should_use_client_side_output_generation(
        client_side_output_generation
    )
    parse_method = resolve_parse_method(file_path, is_ocr, backend)
    run_root, extract_root, archive_zip_path = create_gradio_run_paths(file_path)
    run_root.mkdir(parents=True, exist_ok=True)

    form_data = _api_client.build_parse_request_form_data(
        lang_list=[normalized_language],
        backend=backend,
        effort=effort,
        parse_method=parse_method,
        formula_enable=formula_enable,
        table_enable=table_enable,
        image_analysis=image_analysis,
        server_url=url,
        start_page_id=0,
        end_page_id=end_pages - 1,
        return_md=not use_client_side_output_generation,
        return_middle_json=True,
        return_model_output=True,
        return_content_list=not use_client_side_output_generation,
        return_images=True,
        response_format_zip=True,
        return_original_file=True,
        client_side_output_generation=use_client_side_output_generation,
        enable_idp=enable_idp,
        idp_document_type=idp_document_type,
    )
    upload_assets = [
        _api_client.UploadAsset(
            path=Path(file_path),
            upload_name=build_gradio_upload_name(file_path),
        )
    ]

    async with httpx.AsyncClient(
        timeout=_api_client.build_http_timeout(),
        follow_redirects=True,
    ) as http_client:
        emit_status(STATUS_PREPARING_REQUEST)
        emit_status(STATUS_CHECKING_SERVER)
        server_health = await resolve_server_health(http_client, api_url)
        effective_max_concurrent_requests = resolve_gradio_max_concurrent_requests(
            api_url=api_url,
            server_health=server_health,
        )
        async with _gradio_request_concurrency_limiter.acquire(
            effective_max_concurrent_requests,
            on_wait=lambda snapshot: emit_status(
                format_concurrency_wait_message(snapshot)
            ),
        ):
            emit_status(STATUS_SUBMITTING_TASK)
            submit_response = await _api_client.submit_parse_task(
                base_url=server_health.base_url,
                upload_assets=upload_assets,
                form_data=form_data,
            )
            emit_status(f"Task submitted\uff1atask_id={submit_response.task_id}")

            last_task_snapshot = None

            def handle_task_status(
                status_snapshot: _api_client.TaskStatusSnapshot,
            ) -> None:
                nonlocal last_task_snapshot
                if status_snapshot == last_task_snapshot:
                    return
                last_task_snapshot = status_snapshot
                emit_status(format_remote_status_message(status_snapshot))

            await _api_client.wait_for_task_result(
                client=http_client,
                submit_response=submit_response,
                task_label=Path(file_path).name,
                status_snapshot_callback=handle_task_status,
            )
            emit_status(STATUS_DOWNLOADING_RESULT)
            result_zip_path = await _api_client.download_result_zip(
                client=http_client,
                submit_response=submit_response,
                task_label=Path(file_path).name,
            )

    try:
        _api_client.safe_extract_zip(result_zip_path, extract_root)
    finally:
        result_zip_path.unlink(missing_ok=True)

    file_name = resolve_result_file_name(submit_response, extract_root, file_path)
    local_md_dir = resolve_parse_dir(
        extract_root,
        file_name,
        backend,
        parse_method,
        allow_office_fallback=True,
    )
    emit_status(STATUS_PROCESSING_OUTPUT)
    if use_client_side_output_generation:
        await asyncio.to_thread(regenerate_client_side_outputs, local_md_dir, file_name)

    preview_pdf_path = maybe_generate_local_preview(
        extract_root=extract_root,
        file_name=file_name,
        file_suffix=file_suffix,
        backend=backend,
        parse_method=parse_method,
    )

    zip_archive_success = compress_directory_to_zip(local_md_dir, archive_zip_path)
    if zip_archive_success == 0:
        logger.info('Compression successful')
    else:
        logger.error('Compression failed')

    md_path = Path(local_md_dir) / f"{file_name}.md"
    with open(md_path, 'r', encoding='utf-8') as f:
        txt_content = f.read()
    md_content = replace_image_with_gradio_file_urls(txt_content, local_md_dir)
    content_list_json = read_gradio_content_list_json(local_md_dir, file_name)
    idp_result_json = read_gradio_idp_json(local_md_dir, file_name)

    if file_suffix in office_suffixes:
        preview_pdf_path = None

    emit_status(STATUS_COMPLETED)
    return (
        md_content,
        txt_content,
        content_list_json,
        idp_result_json,
        str(archive_zip_path),
        preview_pdf_path,
    )


async def stream_to_markdown(
    file_path,
    end_pages=10,
    is_ocr=False,
    formula_enable=True,
    table_enable=True,
    image_analysis=True,
    effort=DEFAULT_HYBRID_EFFORT,
    language="ch",
    backend="pipeline",
    url=None,
    api_url=None,
    client_side_output_generation=False,
    enable_idp=False,
    idp_document_type="auto",
):
    status_state = StatusPanelState()
    job_task: asyncio.Task | None = None
    queue_get_task: asyncio.Task | None = None
    timer_task: asyncio.Task | None = None
    yield status_state.render(), None, "", "", "", "", gr.skip()

    if file_path is None:
        return

    status_queue: asyncio.Queue[str] = asyncio.Queue()
    loop = asyncio.get_running_loop()

    def enqueue_status(message: str) -> None:
        loop.call_soon_threadsafe(status_queue.put_nowait, message)

    try:
        job_task = asyncio.create_task(
            _run_to_markdown_job(
                file_path=file_path,
                end_pages=end_pages,
                is_ocr=is_ocr,
                formula_enable=formula_enable,
                table_enable=table_enable,
                image_analysis=image_analysis,
                effort=effort,
                language=language,
                backend=backend,
                url=url,
                api_url=api_url,
                client_side_output_generation=client_side_output_generation,
                enable_idp=enable_idp,
                idp_document_type=idp_document_type,
                status_callback=enqueue_status,
            )
        )

        while True:
            if job_task.done() and status_queue.empty():
                status_state.finalize_processing()
                status_state.finalize_queue()
                break

            queue_get_task = asyncio.create_task(status_queue.get())
            wait_tasks: set[asyncio.Task] = {job_task, queue_get_task}
            timer_task = None
            animation_interval = status_state.animation_interval_seconds
            if animation_interval is not None:
                timer_task = asyncio.create_task(
                    asyncio.sleep(animation_interval)
                )
                wait_tasks.add(timer_task)

            done, pending = await asyncio.wait(
                wait_tasks,
                return_when=asyncio.FIRST_COMPLETED,
            )

            if queue_get_task in done:
                message = queue_get_task.result()
                if status_state.append(message):
                    yield status_state.render(), None, "", "", "", "", gr.skip()
            elif timer_task is not None and timer_task in done:
                if status_state.tick():
                    yield status_state.render(), None, "", "", "", "", gr.skip()
            else:
                queue_get_task.cancel()
                await asyncio.gather(queue_get_task, return_exceptions=True)

            for pending_task in pending:
                if pending_task is job_task:
                    continue
                pending_task.cancel()
                await asyncio.gather(pending_task, return_exceptions=True)
            queue_get_task = None
            timer_task = None

        while not status_queue.empty():
            status_state.append(status_queue.get_nowait())
    except Exception as exc:
        status_state.append(format_failed_status(exc))
        yield status_state.render(), None, "", "", "", "", gr.skip()
        raise
    finally:
        for task in (queue_get_task, timer_task, job_task):
            if task is None or task.done():
                continue
            task.cancel()
            await asyncio.gather(task, return_exceptions=True)

    try:
        (
            md_content,
            txt_content,
            content_list_json,
            idp_result_json,
            archive_zip_path,
            preview_pdf_path,
        ) = await job_task
    except Exception as exc:
        status_state.append(format_failed_status(exc))
        yield status_state.render(), None, "", "", "", "", gr.skip()
        raise

    status_state.append(STATUS_COMPLETED)
    yield (
        status_state.render(),
        archive_zip_path,
        md_content,
        txt_content,
        content_list_json,
        idp_result_json,
        preview_pdf_path,
    )


def resolve_preview_pdf_path(local_md_dir, file_name):
    layout_pdf_path = os.path.join(local_md_dir, file_name + '_layout.pdf')
    if os.path.exists(layout_pdf_path):
        return layout_pdf_path

    origin_pdf_path = os.path.join(local_md_dir, file_name + '_origin.pdf')
    if os.path.exists(origin_pdf_path):
        logger.warning(
            f"Layout preview PDF not found for {file_name}, "
            f"falling back to origin PDF: {origin_pdf_path}"
        )
        return origin_pdf_path

    logger.warning(f"No preview PDF found for {file_name} under {local_md_dir}")
    return None


latex_delimiters_type_a = [
    {'left': '$$', 'right': '$$', 'display': True},
    {'left': '$', 'right': '$', 'display': False},
]
latex_delimiters_type_b = [
    {'left': '\\(', 'right': '\\)', 'display': False},
    {'left': '\\[', 'right': '\\]', 'display': True},
]
latex_delimiters_type_all = latex_delimiters_type_a + latex_delimiters_type_b

header_template = load_resource_text('gradio_header.html')

HEADER_I18N_PLACEHOLDERS = {
    "{{HEADER_TITLE}}": "header_title",
    "{{HEADER_SUBTITLE}}": "header_subtitle",
}
HEADER_GRADIO_VERSION_CLASS_PLACEHOLDER = "{{HEADER_GRADIO_VERSION_CLASS}}"


def render_header_html(i18n):
    """Render the localized application header."""
    rendered_header = header_template
    for placeholder, translation_key in HEADER_I18N_PLACEHOLDERS.items():
        replacement = render_client_i18n_text(i18n, translation_key)
        rendered_header = rendered_header.replace(
            placeholder,
            replacement,
        )
    rendered_header = rendered_header.replace(
        HEADER_GRADIO_VERSION_CLASS_PLACEHOLDER,
        " vsf-gradio6-header" if IS_GRADIO_6 else "",
    )
    return rendered_header

all_lang = [
    ("Trung, Anh, Nhật và chữ Latin (phù hợp tiếng Việt)", "ch"),
    ("Trung, Anh, Nhật và chữ Latin - bản máy chủ", "ch_server"),
    ("Hàn Quốc và tiếng Anh", "korean"),
    ("Tamil và tiếng Anh", "ta"),
    ("Telugu và tiếng Anh", "te"),
    ("Kannada", "ka"),
    ("Thái Lan và tiếng Anh", "th"),
    ("Hy Lạp và tiếng Anh", "el"),
    ("Ả Rập, Ba Tư, Urdu và tiếng Anh", "arabic"),
    ("Nga, Belarus, Ukraina và tiếng Anh", "east_slavic"),
    ("Nhóm chữ Kirin và tiếng Anh", "cyrillic"),
    ("Nhóm chữ Devanagari và tiếng Anh", "devanagari"),
]

hybrid_effort_choices_vi = [
    ("Trung bình - nhanh hơn", "medium"),
    ("Cao - chính xác hơn", "high"),
]


def safe_stem(file_path):
    stem = Path(file_path).stem
    # Implementation detail.
    return re.sub(r'[^\w.]', '_', stem)


def to_pdf(file_path):

    if file_path is None:
        return None

    pdf_bytes = read_fn(file_path)

    # unique_filename = f'{uuid.uuid4()}.pdf'
    unique_filename = f'{safe_stem(file_path)}.pdf'

    # Build the required output.
    tmp_file_path = os.path.join(os.path.dirname(file_path), unique_filename)

    # Process the file path.
    with open(tmp_file_path, 'wb') as tmp_pdf_file:
        tmp_pdf_file.write(pdf_bytes)

    return tmp_file_path


def to_pdf_preview(file_path):
    """Convert the value to the required format."""
    if file_path is None:
        return None
    file_suffix = Path(file_path).suffix.lower().lstrip('.')
    if file_suffix == "pptx":
        return convert_pptx_to_pdf_preview(file_path)
    if file_suffix in office_suffixes:
        return None
    return to_pdf(file_path)


def convert_pptx_to_pdf_preview(file_path):
    """Convert a PPTX upload to a locally served PDF preview."""
    source_path = Path(file_path).resolve()
    libreoffice_path = shutil.which("libreoffice") or shutil.which("soffice")
    if libreoffice_path is None:
        raise RuntimeError("Máy chưa cài LibreOffice để xem trước PPTX.")

    output_dir = source_path.parent / "vsf_pptx_preview"
    output_dir.mkdir(parents=True, exist_ok=True)
    preview_path = output_dir / f"{source_path.stem}.pdf"
    if (
        preview_path.is_file()
        and preview_path.stat().st_mtime >= source_path.stat().st_mtime
    ):
        return str(preview_path)

    profile_dir = Path(tempfile.mkdtemp(prefix="vsf_libreoffice_profile_"))
    try:
        completed = subprocess.run(
            [
                libreoffice_path,
                f"-env:UserInstallation={profile_dir.as_uri()}",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(output_dir),
                str(source_path),
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=60,
        )
    except Exception:
        preview_path.unlink(missing_ok=True)
        raise
    finally:
        shutil.rmtree(profile_dir, ignore_errors=True)

    if completed.returncode != 0 or not preview_path.is_file():
        error_text = (completed.stderr or completed.stdout or "").strip()
        preview_path.unlink(missing_ok=True)
        raise RuntimeError(
            "Không thể chuyển PPTX sang PDF"
            + (f": {error_text}" if error_text else ".")
        )
    return str(preview_path)


def build_gradio_file_public_url(file_path, request: gr.Request):
    """Build the required output."""
    headers = getattr(request, "headers", None) or {}
    host = (
        headers.get('x-forwarded-host')
        or headers.get('host', 'localhost:7860')
    )
    proto = headers.get('x-forwarded-proto', 'http')
    return f"{proto}://{host}/gradio_api/file={quote(str(file_path), safe='/:')}"


def build_short_gradio_file_url(public_url, file_path):
    """Build the required output."""
    base_url = public_url.split("/gradio_api/file=", 1)[0]
    file_name = Path(file_path).name
    file_suffix = Path(file_name).suffix
    file_stem = Path(file_name).stem
    short_file_name = f"{file_stem[-12:]}{file_suffix}" if file_stem else file_name
    return f"{base_url}/....{short_file_name}"


def build_local_xlsx_preview_html(
    file_path,
    i18n=None,
    max_rows=100,
    max_columns=30,
):
    """Render a safe, bounded XLSX preview without an external Office service."""
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        if worksheet.max_row is None or worksheet.max_column is None:
            worksheet.calculate_dimension(force=True)
        worksheet_max_row = worksheet.max_row or 1
        worksheet_max_column = worksheet.max_column or 1
        row_limit = min(max(worksheet_max_row, 1), max_rows)
        column_limit = min(max(worksheet_max_column, 1), max_columns)
        rows = list(
            worksheet.iter_rows(
                min_row=1,
                max_row=row_limit,
                min_col=1,
                max_col=column_limit,
                values_only=True,
            )
        )
        while rows and not any(value not in (None, "") for value in rows[-1]):
            rows.pop()

        sheet_names = "".join(
            (
                '<span class="office-local-sheet'
                + (" is-active" if name == worksheet.title else "")
                + f'">{html_lib.escape(str(name))}</span>'
            )
            for name in workbook.sheetnames
        )
        table_rows = []
        for row_number, row in enumerate(rows, start=1):
            cells = "".join(
                f"<td>{html_lib.escape(str(value)) if value is not None else ''}</td>"
                for value in row
            )
            table_rows.append(f"<tr><th>{row_number}</th>{cells}</tr>")

        is_truncated = (
            worksheet_max_row > max_rows or worksheet_max_column > max_columns
        )
        detail = (
            f"Hiển thị tối đa {max_rows} dòng và {max_columns} cột đầu tiên."
            if is_truncated
            else "Dữ liệu được đọc trực tiếp trên máy, không gửi tới Microsoft."
        )
        body = (
            '<div class="office-local-empty">Bảng tính không có dữ liệu.</div>'
            if not table_rows
            else (
                '<div class="office-local-table-wrap"><table class="office-local-table">'
                "<tbody>"
                + "".join(table_rows)
                + "</tbody></table></div>"
            )
        )
        return (
            '<div class="office-preview-shell office-local-preview">'
            '<div class="office-local-header">'
            "<strong>Xem trước bảng tính cục bộ</strong>"
            f"<span>{html_lib.escape(detail)}</span>"
            f'<div class="office-local-sheets">{sheet_names}</div>'
            "</div>"
            f"{body}"
            "</div>"
        )
    finally:
        workbook.close()


def request_uses_local_host(request: gr.Request) -> bool:
    headers = getattr(request, "headers", None) or {}
    host = (headers.get("x-forwarded-host") or headers.get("host", "")).lower()
    hostname = host.rsplit(":", 1)[0].strip("[]")
    return hostname in {"127.0.0.1", "localhost", "0.0.0.0", "::1"}


def build_local_office_notice_html(file_path):
    file_type = Path(file_path).suffix.lstrip(".").upper() or "Office"
    if file_type == "PPTX":
        title = "Đang chuẩn bị bản xem trước PPTX"
        detail = (
            "Slide được chuyển sang PDF ngay trên máy, không gửi tới Microsoft. "
            "Nếu không chuyển đổi được, tệp vẫn có thể xử lý bình thường."
        )
    else:
        title = f"Không hỗ trợ xem trước {file_type} cục bộ"
        detail = (
            'Tệp vẫn có thể được xử lý bình thường. Nhấn "Bắt đầu xử lý" '
            "để xem nội dung và kết quả."
        )
    return (
        '<div class="office-preview-shell office-local-preview">'
        '<div class="office-local-header">'
        f"<strong>{html_lib.escape(title)}</strong>"
        f"<span>{html_lib.escape(detail)}</span>"
        "</div>"
        "</div>"
    )


def build_pptx_loading_html():
    """Show immediate feedback while LibreOffice creates the PDF preview."""
    return (
        '<div class="office-preview-shell office-local-preview pptx-loading-shell">'
        '<div class="pptx-preview-loading" role="status" aria-live="polite">'
        '<span class="pptx-preview-spinner" aria-hidden="true"></span>'
        '<div class="pptx-preview-loading-copy">'
        "<strong>Đang tạo bản xem trước PPTX…</strong>"
        "<span>Hệ thống đang chuyển slide sang PDF ngay trên máy. "
        "Tệp lớn có thể cần thêm vài giây.</span>"
        "</div>"
        "</div>"
        "</div>"
    )


def build_office_preview_html(file_path, request: gr.Request, i18n=None):
    """Build the required output."""
    file_suffix = Path(file_path).suffix.lower()
    if file_suffix == ".xlsx":
        try:
            return build_local_xlsx_preview_html(file_path, i18n)
        except Exception as exc:
            logger.warning(f"Không thể tạo bản xem trước XLSX cục bộ: {exc}")
            return build_local_office_notice_html(file_path)
    if request_uses_local_host(request):
        return build_local_office_notice_html(file_path)

    public_url = build_gradio_file_public_url(file_path, request)
    short_public_url = build_short_gradio_file_url(public_url, file_path)
    viewer_url = (
        "https://view.officeapps.live.com/op/embed.aspx?src="
        f"{quote(public_url, safe='')}"
    )
    title = render_client_i18n_text(i18n, "office_preview_title")
    notice = render_client_i18n_text(i18n, "office_preview_notice")
    source_link = render_client_i18n_text(i18n, "office_preview_source_link")
    ignore_once = render_client_i18n_text(i18n, "office_preview_ignore_once")
    ignore_forever = render_client_i18n_text(i18n, "office_preview_ignore_forever")
    return (
        '<div class="office-preview-shell">'
        '<div class="office-preview-notice">'
        '<div class="office-preview-copy">'
        f"<strong>{title}</strong>"
        f"<span>{notice}</span>"
        '<div class="office-preview-source-link">'
        f"{source_link}: "
        f"{html_lib.escape(short_public_url)}"
        "</div>"
        "</div>"
        '<div class="office-preview-actions">'
        f'<button type="button" class="office-preview-ignore-once">{ignore_once}</button>'
        f'<button type="button" class="office-preview-ignore-forever">{ignore_forever}</button>'
        "</div>"
        "</div>"
        f'<iframe class="office-preview-frame" src="{html_lib.escape(viewer_url)}" '
        'frameborder="0"></iframe>'
        "</div>"
    )


def update_file_options_html(file_path, request: gr.Request, i18n=None):
    """Process the file path.
    Implementation detail.
    Process the current item.
    Implementation detail.
    """
    if file_path is None:
        return (
            gr.update(visible=True),
            gr.update(value="", visible=False),
            gr.update(value=None, visible=True),
        )

    file_suffix = Path(file_path).suffix.lower().lstrip('.')
    is_office = file_suffix in office_suffixes

    if file_suffix == "pptx":
        return (
            gr.update(visible=True),
            gr.update(value=build_pptx_loading_html(), visible=True),
            gr.update(value=None, visible=False),
        )
    if is_office:
        html_content = build_office_preview_html(file_path, request, i18n)
        return (
            gr.update(visible=True),
            gr.update(value=html_content, visible=True),
            gr.update(value=None, visible=False),
        )
    else:
        return (
            gr.update(visible=True),
            gr.update(value="", visible=False),
            gr.update(visible=True),
        )


def update_doc_show(file_path, request: gr.Request | None = None, i18n=None):
    """Process the file path.
    Process the file path.
    Implementation detail.
    """
    if file_path is None:
        # Process the file path.
        return (
            gr.update(value=None, visible=True),
            gr.update(value="", visible=False),
        )

    file_suffix = Path(file_path).suffix.lower().lstrip('.')

    if file_suffix == "pptx":
        try:
            pdf_path = to_pdf_preview(file_path)
        except Exception as exc:
            logger.warning(f"Không thể tạo bản xem trước PPTX cục bộ: {exc}")
            fallback_html = (
                build_office_preview_html(file_path, request, i18n)
                if request is not None
                else build_local_office_notice_html(file_path)
            )
            return (
                gr.update(value=None, visible=False),
                gr.update(value=fallback_html, visible=True),
            )
        return (
            gr.update(value=pdf_path, visible=True),
            gr.update(value="", visible=False),
        )
    if file_suffix in office_suffixes:
        return gr.update(value=None, visible=False), gr.update()
    else:
        pdf_path = to_pdf_preview(file_path)
        return (
            gr.update(value=pdf_path, visible=True),
            gr.update(value="", visible=False),
        )


@click.command(context_settings=dict(ignore_unknown_options=True, allow_extra_args=True))
@click.pass_context
@click.option(
    '--enable-example',
    'example_enable',
    type=bool,
    help="Enable example files for input."
         "The example files to be input need to be placed in the `examples` folder within the directory where the command is currently executed.",
    default=True,
)
@click.option(
    '--enable-http-client',
    'http_client_enable',
    type=bool,
    help="Enable http-client backend to link openai-compatible servers.",
    default=False,
)
@click.option(
    '--enable-api',
    'api_enable',
    type=bool,
    help="Enable gradio API for serving the application.",
    default=True,
)
@click.option(
    '--max-convert-pages',
    'max_convert_pages',
    type=int,
    help="Set the maximum number of pages to convert from PDF to Markdown.",
    default=1000,
)
@click.option(
    '--server-name',
    'server_name',
    type=str,
    help="Set the server name for the Gradio app.",
    default=None,
)
@click.option(
    '--server-port',
    'server_port',
    type=int,
    help="Set the server port for the Gradio app.",
    default=None,
)
@click.option(
    '--api-url',
    'api_url',
    type=str,
    help="VSF FastAPI base URL. If omitted, gradio starts a reusable local vsf-api service.",
    default=None,
)
@click.option(
    '--enable-vlm-preload',
    'enable_vlm_preload',
    type=bool,
    help="Preload the local VLM model when gradio starts a local vsf-api service.",
    default=False,
)
@click.option(
    '--client-side-output-generation',
    'client_side_output_generation',
    type=bool,
    help="Generate markdown and content lists locally from server-returned middle json.",
    default=False,
)
@click.option(
    '--latex-delimiters-type',
    'latex_delimiters_type',
    type=click.Choice(['a', 'b', 'all']),
    help="Set the type of LaTeX delimiters to use in Markdown rendering:"
         "'a' for type '$', 'b' for type '()[]', 'all' for both types.",
    default='all',
)
def main(ctx,
        example_enable,
        http_client_enable,
        api_enable, max_convert_pages,
        server_name, server_port, api_url, enable_vlm_preload,
        client_side_output_generation, latex_delimiters_type, **kwargs
):

    # Build the required output.
    i18n = gr.I18n(
        en={
            "upload_file": "Chọn hoặc dán tệp cần tải lên\nPDF, hình ảnh, PPTX hoặc XLSX",
            "header_title": "VSF OCR",
            "header_subtitle": "Phân tích tài liệu thông minh, nhận diện văn bản, bố cục, bảng, hình ảnh và công thức.",
            "header_support_text": "Nếu dự án hữu ích, hãy tặng một ⭐️ để ủng hộ!",
            "header_stars_alt": "số sao",
            "header_code_link": "Mã nguồn",
            "header_model_link": "Mô hình",
            "header_model_huggingface_link": "Hugging Face",
            "header_model_modelscope_link": "ModelScope",
            "header_paper_link": "Bài báo",
            "header_paper_vsf_report": "VSF \u00b7 arXiv: 2409.18839",
            "header_paper_vsf25_report": "VSF 2.5 \u00b7 arXiv: 2509.22186",
            "header_paper_vsf25pro_report": "VSF 2.5 Pro \u00b7 arXiv: 2604.04771",
            "header_homepage_link": "Trang chủ",
            "header_download_link": "Tải xuống",
            "max_pages": "Số trang tối đa",
            "backend": "Phương thức xử lý",
            "backend_label_hybrid": "Kết hợp Hybrid (Khuyến nghị)",
            "backend_label_pipeline": "Pipeline (Ổn định, đa ngôn ngữ)",
            "backend_label_vlm": "VLM (Độ chính xác cao cho Trung/Anh)",
            "backend_label_remote_vlm": "VLM từ máy chủ từ xa",
            "backend_label_remote_hybrid": "Hybrid từ máy chủ từ xa",
            "server_url": "Địa chỉ máy chủ",
            "server_url_info": "Địa chỉ máy chủ tương thích OpenAI dành cho chế độ xử lý từ xa.",
            "recognition_options": "**Tùy chọn nhận dạng:**",
            "advanced_options": "Tùy chọn nâng cao",
            "table_enable": "Nhận dạng bảng",
            "table_info": "Nếu tắt, bảng sẽ được giữ dưới dạng hình ảnh.",
            "image_analysis_enable": "Phân tích hình ảnh và biểu đồ",
            "image_analysis_info": "Nếu tắt, hệ thống giữ vị trí ảnh/biểu đồ nhưng không dùng VLM để phân tích nội dung.",
            "formula_label_vlm": "Nhận dạng công thức dạng khối",
            "formula_label_pipeline": "Nhận dạng công thức",
            "formula_label_hybrid": "Nhận dạng công thức trong dòng",
            "formula_info_vlm": "Nếu tắt, công thức dạng khối sẽ được giữ dưới dạng hình ảnh.",
            "formula_info_pipeline": "Nếu tắt, công thức dạng khối sẽ thành hình ảnh và công thức trong dòng sẽ không được nhận dạng.",
            "formula_info_hybrid": "Nếu tắt, công thức trong dòng sẽ không được nhận dạng.",
            "ocr_language": "Ngôn ngữ OCR",
            "ocr_language_info": "Chọn nhóm ngôn ngữ dùng để đọc PDF scan và hình ảnh.",
            "force_ocr": "Buộc chạy OCR",
            "force_ocr_info": "Chỉ bật khi kết quả đọc tự động quá kém; cần chọn đúng ngôn ngữ OCR.",
            "force_ocr_info_hybrid": "Chỉ bật khi kết quả đọc tự động quá kém.",
            "convert": "Bắt đầu xử lý",
            "clear": "Xóa",
            "doc_preview": "Xem trước tài liệu",
            "examples": "Tài liệu mẫu:",
            "convert_status": "Trạng thái xử lý",
            "convert_result": "Tải kết quả",
            "result_file": "Tệp kết quả",
            "md_rendering": "Xem nội dung",
            "md_text": "Markdown",
            "content_list_json": "JSON nội dung",
            "idp_result_json": "Kết quả IDP",
            "status_idle_title": "Đang chờ",
            "status_idle_hint": "Hãy tải tài liệu lên và bắt đầu xử lý.",
            "status_latest": "Trạng thái mới nhất",
            "status_step_prepare": "Chuẩn bị",
            "status_step_check": "Kiểm tra dịch vụ",
            "status_step_submit": "Gửi tác vụ",
            "status_step_queue": "Xếp hàng",
            "status_step_process": "Phân tích",
            "status_step_download": "Tải kết quả",
            "status_step_outputs": "Tạo đầu ra",
            "status_step_done": "Hoàn thành",
            "status_step_failed": "Thất bại",
            "office_preview_title": "Xem trước tài liệu Office",
            "office_preview_notice": "Tính năng xem trước cần tệp có thể truy cập qua Microsoft Office Online và không ảnh hưởng đến quá trình xử lý.",
            "office_preview_source_link": "Đường dẫn tệp",
            "office_preview_ignore_once": "Bỏ qua",
            "office_preview_ignore_forever": "Luôn bỏ qua",
            "backend_info_vlm": "VLM phân tích toàn bộ trang với độ chính xác cao.",
            "backend_info_pipeline": "Pipeline nhiều mô hình, ít tốn tài nguyên và hạn chế sinh sai nội dung.",
            "backend_info_hybrid": "Kết hợp Pipeline và VLM để ưu tiên độ chính xác.",
            "backend_info_default": "Chọn phương thức dùng để phân tích tài liệu.",
            "hybrid_effort": "Mức xử lý Hybrid",
            "hybrid_effort_info": "Trung bình nhanh hơn; Cao chính xác hơn nhưng xử lý lâu hơn.",
            "idp_enable": "Bật IDP tài liệu",
            "idp_info": "Nhận mọi PDF, XLSX, PPTX và ảnh; tự phân loại, trích xuất và kiểm tra dữ liệu.",
            "idp_document_type": "Mẫu dữ liệu cần trích xuất",
            "idp_document_type_info": "Chọn Tự động để nhận mọi tài liệu. Các lựa chọn bên dưới chỉ là mẫu trích xuất chuyên sâu.",
        },
        zh={
            "upload_file": "Chọn hoặc dán tệp cần tải lên\nPDF, hình ảnh, PPTX hoặc XLSX",
            "header_title": "VSF OCR",
            "header_subtitle": "Phân tích tài liệu thông minh, nhận diện văn bản, bố cục, bảng, hình ảnh và công thức.",
            "header_support_text": "Nếu dự án hữu ích, hãy tặng một ⭐️ để ủng hộ!",
            "header_stars_alt": "số sao",
            "header_code_link": "Mã nguồn",
            "header_model_link": "Mô hình",
            "header_model_huggingface_link": "Hugging Face",
            "header_model_modelscope_link": "ModelScope",
            "header_paper_link": "Bài báo",
            "header_paper_vsf_report": "VSF \u00b7 arXiv: 2409.18839",
            "header_paper_vsf25_report": "VSF 2.5 \u00b7 arXiv: 2509.22186",
            "header_paper_vsf25pro_report": "VSF 2.5 Pro \u00b7 arXiv: 2604.04771",
            "header_homepage_link": "Trang chủ",
            "header_download_link": "Tải xuống",
            "max_pages": "Số trang tối đa",
            "backend": "Phương thức xử lý",
            "backend_label_hybrid": "Kết hợp Hybrid (Khuyến nghị)",
            "backend_label_pipeline": "Pipeline (Ổn định, đa ngôn ngữ)",
            "backend_label_vlm": "VLM (Độ chính xác cao cho Trung/Anh)",
            "backend_label_remote_vlm": "VLM từ máy chủ từ xa",
            "backend_label_remote_hybrid": "Hybrid từ máy chủ từ xa",
            "server_url": "Địa chỉ máy chủ",
            "server_url_info": "Địa chỉ máy chủ tương thích OpenAI dành cho chế độ xử lý từ xa.",
            "recognition_options": "**Tùy chọn nhận dạng:**",
            "advanced_options": "Tùy chọn nâng cao",
            "table_enable": "Nhận dạng bảng",
            "table_info": "Nếu tắt, bảng sẽ được giữ dưới dạng hình ảnh.",
            "image_analysis_enable": "Phân tích hình ảnh và biểu đồ",
            "image_analysis_info": "Nếu tắt, hệ thống giữ vị trí ảnh/biểu đồ nhưng không dùng VLM để phân tích nội dung.",
            "formula_label_vlm": "Nhận dạng công thức dạng khối",
            "formula_label_pipeline": "Nhận dạng công thức",
            "formula_label_hybrid": "Nhận dạng công thức trong dòng",
            "formula_info_vlm": "Nếu tắt, công thức dạng khối sẽ được giữ dưới dạng hình ảnh.",
            "formula_info_pipeline": "Nếu tắt, công thức dạng khối sẽ thành hình ảnh và công thức trong dòng sẽ không được nhận dạng.",
            "formula_info_hybrid": "Nếu tắt, công thức trong dòng sẽ không được nhận dạng.",
            "ocr_language": "Ngôn ngữ OCR",
            "ocr_language_info": "Chọn nhóm ngôn ngữ dùng để đọc PDF scan và hình ảnh.",
            "force_ocr": "Buộc chạy OCR",
            "force_ocr_info": "Chỉ bật khi kết quả đọc tự động quá kém; cần chọn đúng ngôn ngữ OCR.",
            "force_ocr_info_hybrid": "Chỉ bật khi kết quả đọc tự động quá kém.",
            "convert": "Bắt đầu xử lý",
            "clear": "Xóa",
            "doc_preview": "Xem trước tài liệu",
            "examples": "Tài liệu mẫu:",
            "convert_status": "Trạng thái xử lý",
            "convert_result": "Tải kết quả",
            "result_file": "Tệp kết quả",
            "md_rendering": "Xem nội dung",
            "md_text": "Markdown",
            "content_list_json": "JSON nội dung",
            "idp_result_json": "Kết quả IDP",
            "status_idle_title": "Đang chờ",
            "status_idle_hint": "Hãy tải tài liệu lên và bắt đầu xử lý.",
            "status_latest": "Trạng thái mới nhất",
            "status_step_prepare": "Chuẩn bị",
            "status_step_check": "Kiểm tra dịch vụ",
            "status_step_submit": "Gửi tác vụ",
            "status_step_queue": "Xếp hàng",
            "status_step_process": "Phân tích",
            "status_step_download": "Tải kết quả",
            "status_step_outputs": "Tạo đầu ra",
            "status_step_done": "Hoàn thành",
            "status_step_failed": "Thất bại",
            "office_preview_title": "Xem trước tài liệu Office",
            "office_preview_notice": "Tính năng xem trước cần tệp có thể truy cập qua Microsoft Office Online và không ảnh hưởng đến quá trình xử lý.",
            "office_preview_source_link": "Đường dẫn tệp",
            "office_preview_ignore_once": "Bỏ qua",
            "office_preview_ignore_forever": "Luôn bỏ qua",
            "backend_info_vlm": "VLM phân tích toàn bộ trang với độ chính xác cao.",
            "backend_info_pipeline": "Pipeline nhiều mô hình, ít tốn tài nguyên và hạn chế sinh sai nội dung.",
            "backend_info_hybrid": "Kết hợp Pipeline và VLM để ưu tiên độ chính xác.",
            "backend_info_default": "Chọn phương thức dùng để phân tích tài liệu.",
            "hybrid_effort": "Mức xử lý Hybrid",
            "hybrid_effort_info": "Trung bình nhanh hơn; Cao chính xác hơn nhưng xử lý lâu hơn.",
            "idp_enable": "Bật IDP tài liệu",
            "idp_info": "Nhận mọi PDF, XLSX, PPTX và ảnh; tự phân loại, trích xuất và kiểm tra dữ liệu.",
            "idp_document_type": "Mẫu dữ liệu cần trích xuất",
            "idp_document_type_info": "Chọn Tự động để nhận mọi tài liệu. Các lựa chọn bên dưới chỉ là mẫu trích xuất chuyên sâu.",
        },
    )

    # Extract the required value.
    def get_formula_label(backend_choice):
        if backend_choice.startswith("vlm"):
            return i18n("formula_label_vlm")
        elif backend_choice == "pipeline":
            return i18n("formula_label_pipeline")
        elif backend_choice.startswith("hybrid"):
            return i18n("formula_label_hybrid")
        else:
            return i18n("formula_label_pipeline")

    def get_formula_info(backend_choice):
        if backend_choice.startswith("vlm"):
            return i18n("formula_info_vlm")
        elif backend_choice == "pipeline":
            return i18n("formula_info_pipeline")
        elif backend_choice.startswith("hybrid"):
            return i18n("formula_info_hybrid")
        else:
            return ""

    def get_backend_info(backend_choice):
        return i18n(select_backend_info_key(backend_choice))

    def get_force_ocr_info(backend_choice):
        """Prepare the output value."""
        return i18n(select_force_ocr_info_key(backend_choice))

    def build_interface_updates(backend_choice, effort_choice):
        """Build the required output."""
        formula_label_update = gr.update(label=get_formula_label(backend_choice), info=get_formula_info(backend_choice))
        backend_info_update = gr.update(info=get_backend_info(backend_choice))
        force_ocr_update = gr.update(info=get_force_ocr_info(backend_choice))

        return (
            force_ocr_update,
            formula_label_update,
            backend_info_update,
        )

    def update_interface(backend_choice, effort_choice):
        """Process the current item."""
        return build_interface_updates(backend_choice, effort_choice)

    del kwargs
    _gradio_local_api_server.configure(
        resolve_gradio_local_api_cli_args(
            ctx.args,
            api_url=api_url,
            enable_vlm_preload=enable_vlm_preload,
        )
    )

    if latex_delimiters_type == 'a':
        latex_delimiters = latex_delimiters_type_a
    elif latex_delimiters_type == 'b':
        latex_delimiters = latex_delimiters_type_b
    elif latex_delimiters_type == 'all':
        latex_delimiters = latex_delimiters_type_all
    else:
        raise ValueError(f"Invalid latex delimiters type: {latex_delimiters_type}.")


    async def convert_to_markdown_stream(
        file_path,
        end_pages=10,
        is_ocr=False,
        formula_enable=True,
        table_enable=True,
        image_analysis=True,
        effort=DEFAULT_HYBRID_EFFORT,
        language="ch",
        backend="pipeline",
        url=None,
        enable_idp=False,
        idp_document_type="auto",
        request: gr.Request = None,
    ):
        request_locale = resolve_request_locale(request)
        async for update in stream_to_markdown(
            file_path=file_path,
            end_pages=end_pages,
            is_ocr=is_ocr,
            formula_enable=formula_enable,
            table_enable=table_enable,
            image_analysis=image_analysis,
            effort=effort,
            language=language,
            backend=backend,
            url=url,
            api_url=api_url,
            client_side_output_generation=client_side_output_generation,
            enable_idp=enable_idp,
            idp_document_type=idp_document_type,
        ):
            update = (
                render_status_steps_html(update[0], i18n, locale=request_locale),
                update[1],
                prepare_markdown_for_gradio_preview(update[2], latex_delimiters),
                *update[3:],
            )
            yield update

    suffixes = [f".{suffix}" for suffix in pdf_suffixes + image_suffixes + office_suffixes]
    _blocks_kwargs = {"title": "VSF OCR"}
    if not IS_GRADIO_6:
        _blocks_kwargs.update({"css": APP_CSS, "js": APP_JS})
    with gr.Blocks(**_blocks_kwargs) as demo:
        gr.HTML(render_header_html(i18n), elem_classes=["vsf-header-html"])
        with gr.Row(elem_classes=["vsf-workspace-row"]):
            with gr.Column(variant='panel', scale=2, min_width=280, elem_classes=["vsf-control-column"]):
                input_file = gr.File(
                    label=i18n("upload_file"),
                    file_types=suffixes,
                    elem_classes=["vsf-upload-file"],
                )
                preferred_option = DEFAULT_BACKEND
                backend = gr.Dropdown(
                    build_backend_choices(http_client_enable, i18n),
                    label=i18n("backend"),
                    value=preferred_option,
                    info=get_backend_info(preferred_option),
                    elem_classes=["vsf-backend-select"],
                )
                with gr.Row(
                    visible=frontend_managed_initial_visibility(is_http_client_backend(preferred_option)),
                    elem_classes=["vsf-client-options"],
                ):
                    url = gr.Textbox(
                        label=i18n("server_url"),
                        value='http://localhost:30000',
                        placeholder='http://localhost:30000',
                        info=i18n("server_url_info"),
                    )
                # Process the file path.
                with gr.Group() as options_group:
                    max_pages = gr.Slider(1, max_convert_pages, max_convert_pages, step=1, label=i18n("max_pages"))
                    gr.Button(
                        i18n("advanced_options"),
                        size="sm",
                        elem_classes=["vsf-advanced-open"],
                    )
                with gr.Row(elem_classes=["vsf-actions"]):
                    change_bu = gr.Button(i18n("convert"), variant="primary", scale=1, min_width=0)
                    clear_bu = gr.ClearButton(value=i18n("clear"), scale=1, min_width=0)
                output_file = gr.File(
                    label=i18n("convert_result"),
                    interactive=False,
                    elem_classes=["vsf-result-file"],
                )
                status_panel = gr.HTML(
                    value=render_status_steps_html("", i18n),
                    label=i18n("convert_status"),
                    elem_classes=["vsf-status-panel"],
                )

            _doc_preview_label = i18n("doc_preview")
            # Implementation detail.
            # Implementation detail.
            preview_content_height = 775
            pdf_preview_page_height = 720
            with gr.Column(variant='panel', scale=4, min_width=340, elem_classes=["vsf-preview-pane"]):
                doc_show = PDF(
                    label=_doc_preview_label,
                    interactive=False,
                    visible=True,
                    height=pdf_preview_page_height,
                )
                office_html = gr.HTML(
                    value="",
                    visible=False,
                    min_height=preview_content_height,
                    elem_classes=["vsf-office-preview-html"],
                )

            with gr.Column(variant='panel', scale=4, min_width=340, elem_classes=["vsf-markdown-pane"]):
                _md_copy_kwargs = {"buttons": ["copy"]} if IS_GRADIO_6 else {"show_copy_button": True}
                _textarea_copy_kwargs = {"buttons": ["copy"]} if IS_GRADIO_6 else {"show_copy_button": True}
                with gr.Tabs(elem_classes=["vsf-markdown-tabs"]):
                    with gr.Tab(i18n("md_rendering")):
                        md = gr.Markdown(
                            label=i18n("md_rendering"),
                            height=preview_content_height,
                            elem_classes=["vsf-markdown-output"],
                            latex_delimiters=latex_delimiters,
                            line_breaks=True,
                            **_md_copy_kwargs
                        )
                    with gr.Tab(i18n("idp_result_json")):
                        idp_result_json = gr.Code(
                            lines=28,
                            language="json",
                            label=i18n("idp_result_json"),
                            interactive=False,
                            wrap_lines=True,
                            show_label=False,
                            elem_classes=["vsf-idp-result-json"],
                        )
                    with gr.Tab(i18n("md_text")):
                        md_text = gr.Code(
                            lines=28,
                            language="markdown",
                            label=i18n("md_text"),
                            interactive=False,
                            wrap_lines=True,
                            show_label=False,
                            elem_classes=["vsf-markdown-text"],
                        )
                    with gr.Tab(i18n("content_list_json")):
                        content_list_json = gr.Code(
                            lines=28,
                            language="json",
                            label=i18n("content_list_json"),
                            interactive=False,
                            wrap_lines=True,
                            show_label=False,
                            elem_classes=["vsf-content-list-json"],
                        )

        if example_enable:
            example_root = os.path.join(os.getcwd(), 'examples')
            if os.path.exists(example_root):
                example_files = [
                    os.path.join(example_root, _) for _ in os.listdir(example_root)
                    if _.endswith(tuple(suffixes))
                ]
                if example_files:
                    with gr.Accordion(i18n("examples"), open=True, elem_classes=["vsf-examples-panel"]):
                        gr.Examples(
                            examples=example_files,
                            inputs=input_file,
                            elem_id="vsf-example-files",
                            label=None,
                        )

        with gr.Column(elem_classes=["vsf-advanced-popover"]):
            with gr.Column(elem_classes=["vsf-advanced-card"]):
                with gr.Group():
                    table_enable = gr.Checkbox(label=i18n("table_enable"), value=True, info=i18n("table_info"))
                    formula_enable = gr.Checkbox(label=get_formula_label(preferred_option), value=True, info=get_formula_info(preferred_option))
                    image_analysis = gr.Checkbox(
                        label=i18n("image_analysis_enable"),
                        value=True,
                        visible=frontend_managed_initial_visibility(
                            is_image_analysis_option_visible(preferred_option, DEFAULT_HYBRID_EFFORT)
                        ),
                        info=i18n("image_analysis_info"),
                        elem_classes=["vsf-image-analysis-option"],
                    )
                    enable_idp = gr.Checkbox(
                        label=i18n("idp_enable"),
                        value=True,
                        info=i18n("idp_info"),
                    )
                    idp_document_type = gr.Dropdown(
                        choices=[
                            ("Tự động – nhận mọi loại tài liệu", "auto"),
                            *[
                                (schema.label, document_type)
                                for document_type, schema in DOCUMENT_SCHEMAS.items()
                            ],
                        ],
                        value="auto",
                        label=i18n("idp_document_type"),
                        info=i18n("idp_document_type_info"),
                    )
                    with gr.Column(elem_classes=["vsf-hybrid-effort-option"]):
                        hybrid_effort = gr.Radio(
                            hybrid_effort_choices_vi,
                            label=i18n("hybrid_effort"),
                            value=DEFAULT_HYBRID_EFFORT,
                            info=i18n("hybrid_effort_info"),
                            elem_classes=["vsf-hybrid-effort"],
                        )
                with gr.Column(elem_classes=["vsf-force-ocr-option"]):
                    with gr.Group():
                        with gr.Column(elem_classes=["vsf-ocr-language-options"]):
                            language = gr.Dropdown(
                                all_lang,
                                label=i18n("ocr_language"),
                                value="ch",
                                info=i18n("ocr_language_info"),
                            )
                        is_ocr = gr.Checkbox(
                            label=i18n("force_ocr"),
                            value=False,
                            info=i18n(select_force_ocr_info_key(preferred_option)),
                        )

        # Add the value to the result.
        _private_api_kwargs = (
            {"api_visibility": "private", "queue": False, "show_progress": "hidden"}
            if IS_GRADIO_6
            else {"api_name": False, "queue": False, "show_progress": "hidden"}
        )
        backend.change(
            fn=update_interface,
            inputs=[backend, hybrid_effort],
            outputs=[is_ocr, formula_enable, backend],
            **_private_api_kwargs
        )
        # Add the value to the result.
        demo.load(
            fn=update_interface,
            inputs=[backend, hybrid_effort],
            outputs=[is_ocr, formula_enable, backend],
            **_private_api_kwargs
        )
        clear_bu.add([
            input_file,
            md,
            doc_show,
            md_text,
            content_list_json,
            idp_result_json,
            output_file,
            is_ocr,
            enable_idp,
            idp_document_type,
            office_html,
            status_panel,
        ])

        def reset_primary_ui():
            """Implementation detail."""
            return (
                gr.update(visible=True),
                gr.update(value=None, visible=True),
                gr.update(value="", visible=False),
                gr.update(value=render_status_steps_html("", i18n)),
                gr.update(value=""),
            )

        # Implementation detail.
        clear_bu.click(
            fn=reset_primary_ui,
            inputs=[],
            outputs=[options_group, doc_show, office_html, status_panel, content_list_json],
            **_private_api_kwargs
        )

        def update_file_options_html_for_ui(file_path, request: gr.Request):
            """Process the file path."""
            return update_file_options_html(file_path, request, i18n)

        def update_doc_show_for_ui(file_path, request: gr.Request):
            """Build one mutually exclusive local or remote preview."""
            return update_doc_show(file_path, request, i18n)

        # Implementation detail.
        # Implementation detail.
        # Implementation detail.
        input_file.change(
            fn=update_file_options_html_for_ui,
            inputs=input_file,
            outputs=[options_group, office_html, doc_show],
            **_private_api_kwargs
        ).then(
            fn=update_doc_show_for_ui,
            inputs=input_file,
            outputs=[doc_show, office_html],
            **_private_api_kwargs
        )
        _to_md_api_kwargs = (
            {
                "api_visibility": "public" if api_enable else "private",
                "queue": True,
                "show_progress": "hidden",
            }
            if IS_GRADIO_6
            else {
                "api_name": "to_markdown" if api_enable else False,
                "queue": True,
                "show_progress": "hidden",
            }
        )
        change_bu.click(
            fn=convert_to_markdown_stream,
            inputs=[
                input_file,
                max_pages,
                is_ocr,
                formula_enable,
                table_enable,
                image_analysis,
                hybrid_effort,
                language,
                backend,
                url,
                enable_idp,
                idp_document_type,
            ],
            outputs=[
                status_panel,
                output_file,
                md,
                md_text,
                content_list_json,
                idp_result_json,
                doc_show,
            ],
            **_to_md_api_kwargs
        )

    demo.queue(default_concurrency_limit=None)

    if IS_GRADIO_6:
        footer_links = ["gradio", "settings"]
        if api_enable:
            footer_links.append("api")
        _launch_kwargs = {"footer_links": footer_links, "css": APP_CSS, "head": APP_HEAD}
    else:
        _launch_kwargs = {"show_api": api_enable}
    maybe_prepare_local_api_for_gradio_startup(
        api_url=api_url,
        enable_vlm_preload=enable_vlm_preload,
    )
    demo.launch(
        server_name=server_name,
        server_port=server_port,
        i18n=i18n,
        allowed_paths=build_gradio_allowed_paths(),
        **_launch_kwargs,
    )


if __name__ == '__main__':
    main()
